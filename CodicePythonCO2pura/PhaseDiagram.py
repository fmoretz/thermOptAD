#---------------------------------------------------
# The main goal of this program is to define the
# phase equilibrium conditions for carbon dioxide
# and its thermo-chemical properties of interest
# at a given pressure and temperature
#---------------------------------------------------

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from scipy import optimize as opt
# Module containing Span-Wagner correlation equations
import SWcorrelationsEq as SWcE
# Module containing the algorithms developed to obtain carbon dioxide density
from densityAlgorithm import density
# Module containing functions developed to obtain carbon dioxide properties
import CO2properties as CO2p

eps  = 1.0E-6

#-----------------------
# data.xlsx format is:        
# n, T[K], P[Pa] 
#-----------------------

print("The algorithms are reading the data file")

# Reading data set (data range is: 73.15 K<T<323.15 K and 0.01 MPa<P<10 MPa)
path = 'PureCO2.xlsx'
data = pd.read_excel(path, sheet_name="Dati", index_col=0)
# Temperature [K] and pressure [Pa] vectors
T_i = data['T[K]'].to_numpy()
P_i = data['P[Pa]'].to_numpy()
# Reading record number
imax = len(data.axes[0])

# results sheet in the PureCO2.xlsx file
results = pd.DataFrame(index=False, columns=['N', "T[K]", "P[MPa]", 'P_eq[MPa]', 'Phase'])
# co2Sproperies sheet in the PureCO2.xlsx file
co2S = pd.DataFrame(index=False, columns=['N', "T[K]", "P[MPa]", "d[kg/m^(3)]", 's[J/(mol*K)]', 'h[J/mol]',\
			   "c_p[J/(mol*K)]", "c_v[J/(mol*K)]","a[K^(-1)]"])
# co2Lproperties sheet in the PureCO2.xlsx file
co2L = pd.DataFrame(index=False,columns=['N', "T[K]", "P[MPa]", "d[kg/m^(3)]", 's[J/(mol*K)]', 'h[J/mol]',\
			   "c_p[J/(mol*K)]", "c_v[J/(mol*K)]","a[K^(-1)]","lambda[W/(m*K)]","eta[Pa*s]","sigma[N/m]"])
# co2Vproperties sheet in the PureCO2.xlsx file
co2V = pd.DataFrame(index=False, columns=['N', "T[K]", "P[MPa]", "d[kg/m^(3)]", 's[J/(mol*K)]', 'h[J/mol]',\
			   "c_p[J/(mol*K)]", "c_v[J/(mol*K)]","a[K^(-1)]","lambda[W/(m*K)]","eta[Pa*s]"])
# co2Cproperties sheet in the PureCO2.xlsx file
co2C = pd.DataFrame(index=False, columns=['N', "T[K]", "P[MPa]", "d[kg/m^(3)]", 's[J/(mol*K)]', 'h[J/mol]',\
			   "c_p[J/(mol*K)]", "c_v[J/(mol*K)]","a[K^(-1)]","lambda[W/(m*K)]","eta[Pa*s]"])

with pd.ExcelWriter(path) as writer:
	writer.book = openpyxl.load_workbook(path)
	results.to_excel(writer, sheet_name='Results')
	co2S.to_excel(writer, sheet_name='co2Sproperies')
	co2L.to_excel(writer, sheet_name='co2Lproperies')
	co2V.to_excel(writer, sheet_name='co2Vproperies')
	co2C.to_excel(writer, sheet_name='co2Cproperies')

# Start loop over i
for i in range(imax-1):
	d   = 0.
	s   = 0.
	h   = 0.
	c_p = 0.
	c_v = 0.
	a   = 0.
	dl  = 0.
	s_l = 0.
	h_l = 0.
	cpl = 0.
	cvl = 0.
	a_l = 0.
	dv  = 0.
	s_v = 0.
	h_v = 0.
	cpv = 0.
	cvv = 0.
	a_v = 0.
	st  = 0.
	tc  = 0.
	v   = 0.
	#-----------------------------------------------------
	# WARNING: P and x in output files are in units of   
	# MPa, while data pressure have to be expressed in   
	# units of Pa in order to obtain the right physical  
	# dimension of the Gibbs free energy (function f(x)) 
	#-----------------------------------------------------
	T   = T_i[i]
	P   = P_i[i]
	P_d = P[i]*1.0E-06
	# Check introduced in order to analyse only record within the data range
	if ((T >= 73.15 and T <= 323.15) and (P >= 0.1E5 and P <= 1.0E7)): 
		if ((abs(T - T_t) < 1.0E-03) and (abs((P - P_t)*1.0E-05) < 1.0E-04)): 
			# Triple point
			dfr = pd.DataFrame([i, T, P_d, P_d, "TP"])
			dfr.style.apply(highlight_max, color='black')
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
    
			# Solid properties
			d   = CO2p.SolidProperties(1,T,P)
			s   = CO2p.SolidProperties(2,T,P)
			h   = CO2p.SolidProperties(3,T,P)
			c_p = CO2p.SolidProperties(4,T,P)
			c_v = CO2p.SolidProperties(5,T,P)
			a   = CO2p.SolidProperties(6,T,P)
			dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
			dfs.style.apply(highlight_max, color='black')
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
			# Liquid properties
			dl  = density(1,T,P)
			s_l = CO2p.VLProperties(1,T,dl)
			h_l = CO2p.VLProperties(2,T,dl)
			cpl = CO2p.VLProperties(3,T,dl)
			cvl = CO2p.VLProperties(4,T,dl)
			a_l = CO2p.VLProperties(5,T,dl)
			tc  = CO2p.VLProperties(6,T,dl)
			v   = CO2p.VLProperties(7,T,dl)
			st  = CO2p.VLProperties(8,T,dl)
			dfl = pd.DataFrame([i, T, P_d, dl, s_l, h_l, cpl, cvl, a_l, tc, v, st])
			dfl.style.apply(highlight_max, color='black')
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfl.to_excel(writer, sheet_name="co2Lproperies", startrow=i+2)
			# Vapor properties
			dv  = density(2,T,P)
			s_v = CO2p.VLProperties(1,T,dv)
			h_v = CO2p.VLProperties(2,T,dv)
			cpv = CO2p.VLProperties(3,T,dv)
			cvv = CO2p.VLProperties(4,T,dv)
			a_v = CO2p.VLProperties(5,T,dv)
			tc  = CO2p.VLProperties(6,T,dv)
			v   = CO2p.VLProperties(7,T,dv)
			dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
			dfv.style.apply(highlight_max, color='black')
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfv.to_excel(writer, sheet_name="co2Vproperies", startrow=i+2)
		##! Did you mean T<= 150.0 only?
		elif(T <= 73.15 and T <= 150.0): 
			# Solid Phase (S)
			dfr = pd.DataFrame([i, T, P_d, "null", "S"])
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
			# Solid properties
			d   = CO2p.SolidProperties(1,T,P)
			s   = CO2p.SolidProperties(2,T,P)
			h   = CO2p.SolidProperties(3,T,P)
			c_p = CO2p.SolidProperties(4,T,P)
			c_v = CO2p.SolidProperties(5,T,P)
			a   = CO2p.SolidProperties(6,T,P)
			dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
			with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
				dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
		elif (T > 150.0 and T <= T_t): 
			#-----------------
			# SV Equilibrium  
			#-----------------
			P_e = (EquilibriumSolver(eps,1,T))*1.0E-06
			if ((abs((P - x)*1.0E-05) > 1.0E-04) and P < x):
				# Vapor Phase (V)
				dfr = pd.DataFrame([i, T, P_d, P_e, "V"])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
				# Vapor properties
				dv  = density(2,T,P)
				s_v = CO2p.VLProperties(1,T,dv)
				h_v = CO2p.VLProperties(2,T,dv)
				cpv = CO2p.VLProperties(3,T,dv)
				cvv = CO2p.VLProperties(4,T,dv)
				a_v = CO2p.VLProperties(5,T,dv)
				tc  = CO2p.VLProperties(6,T,dv)
				v   = CO2p.VLProperties(7,T,dv)
				dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfv.to_excel(writer, sheet_name="co2Vproperies", startrow=i+2)
			elif ((abs((P - x)*1.0E-05) > 1.0E-04) and P > x):
				# Solid Phase (S)
				dfr = pd.DataFrame([i, T, P_d, P_e, "S"])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
				# Solid properties
				d   = CO2p.SolidProperties(1,T,P)
				s   = CO2p.SolidProperties(2,T,P)
				h   = CO2p.SolidProperties(3,T,P)
				c_p = CO2p.SolidProperties(4,T,P)
				c_v = CO2p.SolidProperties(5,T,P)
				a   = CO2p.SolidProperties(6,T,P)
				dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
			else:
				# Sublimation line (SVE)
				dfr = pd.DataFrame([i, T, P_d, P_e, "SVE"])
				dfr.style.apply(highlight_max, color='blue')
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
				# Vapor properties
				dv  = density(2,T,P)
				s_v = CO2p.VLProperties(1,T,dv)
				h_v = CO2p.VLProperties(2,T,dv)
				cpv = CO2p.VLProperties(3,T,dv)
				cvv = CO2p.VLProperties(4,T,dv)
				a_v = CO2p.VLProperties(5,T,dv)
				tc  = CO2p.VLProperties(6,T,dv)
				v   = CO2p.VLProperties(7,T,dv)
				dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
				dfv.style.apply(highlight_max, color='blue')
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfv.to_excel(writer, sheet_name="co2Vproperies", startrow=i+2)
				# Solid properties
				d   = CO2p.SolidProperties(1,T,P)
				s   = CO2p.SolidProperties(2,T,P)
				h   = CO2p.SolidProperties(3,T,P)
				c_p = CO2p.SolidProperties(4,T,P)
				c_v = CO2p.SolidProperties(5,T,P)
				a   = CO2p.SolidProperties(6,T,P)
				dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
				dfs.style.apply(highlight_max, color='blue')
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
		else:
			if ((abs(T - T_c) < 1.0E-03) and (abs((P - P_c)*1.0E-05) < 1.0E-04)):
				# Critical point
				dfr = pd.DataFrame([i, T, P_d, P_d, "CP"])
				dfr.style.apply(highlight_max, color='black')
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
				# Critical point properties
				dv  = density(2,T,P)
				s_v = CO2p.VLProperties(1,T,dv)
				h_v = CO2p.VLProperties(2,T,dv)
				cpv = CO2p.VLProperties(3,T,dv)
				cvv = CO2p.VLProperties(4,T,dv)
				a_v = CO2p.VLProperties(5,T,dv)
				tc  = CO2p.VLProperties(6,T,dv)
				v   = CO2p.VLProperties(7,T,dv)
				dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
				dfv.style.apply(highlight_max, color='black')
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfv.to_excel(writer, sheet_name="co2Cproperies", startrow=i+2)
			elif T >= T_c:
				#------------------
				# Critical region 
				#------------------
				dfr = pd.DataFrame([i, T, P_d, "null", "C"])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
				# Critical properties
				dv  = density(2,T,P)
				s_v = CO2p.VLProperties(1,T,dv)
				h_v = CO2p.VLProperties(2,T,dv)
				cpv = CO2p.VLProperties(3,T,dv)
				cvv = CO2p.VLProperties(4,T,dv)
				a_v = CO2p.VLProperties(5,T,dv)
				tc  = CO2p.VLProperties(6,T,dv)
				v   = CO2p.VLProperties(7,T,dv)
				dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
				with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
					dfv.to_excel(writer, sheet_name="co2Cproperies", startrow=i+2)
			else:
				#-----------------
				# VL Equilibrium  
				#-----------------
				P_e = EquilibriumSolver(eps,2,T)*1.0E-06
				if ((abs((P - x)*1.0E-05) > 1.0E-04) and P < x):
					# Vapor Phase (V)
					dfr = pd.DataFrame([i, T, P_d, P_e, "V"])
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfr.to_excel(writer, sheet_name="Results", startrow=i+2)                   
					# Vapor properties
					dv  = density(2,T,P)
					s_v = CO2p.VLProperties(1,T,dv)
					h_v = CO2p.VLProperties(2,T,dv)
					cpv = CO2p.VLProperties(3,T,dv)
					cvv = CO2p.VLProperties(4,T,dv)
					a_v = CO2p.VLProperties(5,T,dv)
					tc  = CO2p.VLProperties(6,T,dv)
					v   = CO2p.VLProperties(7,T,dv)
					dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfv.to_excel(writer, sheet_name="co2Vproperies", startrow=i+2)
				elif (((abs((P - x)*1.0E-05) > 1.0E-04) and P > x) and (T > 240 and T < T_c)):
					# Liquid Phase (L)
					dfr = pd.DataFrame([i, T, P_d, P_e, "L"])
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
					# Liquid properties
					dl  = density(1,T,P)
					s_l = CO2p.VLProperties(1,T,dl)
					h_l = CO2p.VLProperties(2,T,dl)
					cpl = CO2p.VLProperties(3,T,dl)
					cvl = CO2p.VLProperties(4,T,dl)
					a_l = CO2p.VLProperties(5,T,dl)
					tc  = CO2p.VLProperties(6,T,dl)
					v   = CO2p.VLProperties(7,T,dl)
					st  = CO2p.VLProperties(8,T,dl)
					dfl = pd.DataFrame([i, T, P_d, dl, s_l, h_l, cpl, cvl, a_l, tc, v, st])
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfl.to_excel(writer, sheet_name="co2Lproperies", startrow=i+2)
				elif(((abs((P - x)*1.0E-05) > 1.0E-04) and P > x) and T <= 240):
					#-----------------
					# SL Equilibrium 
					#-----------------
					P_e = EquilibriumSolver(eps,3,T)*1.0E-06
					if ((abs((P - x)*1.0E-05) > 1.0E-04) and P < x):
						# Liquid Phase (L)
						dfr = pd.DataFrame([i, T, P_d, P_e, "L"])
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
						# Liquid properties
						dl  = density(1,T,P)
						s_l = CO2p.VLProperties(1,T,dl)
						h_l = CO2p.VLProperties(2,T,dl)
						cpl = CO2p.VLProperties(3,T,dl)
						cvl = CO2p.VLProperties(4,T,dl)
						a_l = CO2p.VLProperties(5,T,dl)
						tc  = CO2p.VLProperties(6,T,dl)
						v   = CO2p.VLProperties(7,T,dl)
						st  = CO2p.VLProperties(8,T,dl)
						dfl = pd.DataFrame([i, T, P_d, dl, s_l, h_l, cpl, cvl, a_l, tc, v, st])
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfl.to_excel(writer, sheet_name="co2Lproperies", startrow=i+2)
					elif ((abs((P - x)*1.0E-05) > 1.0E-04) and P > x):
						# Solid Phase (S)
						dfr = pd.DataFrame([i, T, P_d, P_e, "S"])
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
						# Solid properties
						d   = CO2p.SolidProperties(1,T,P)
						s   = CO2p.SolidProperties(2,T,P)
						h   = CO2p.SolidProperties(3,T,P)
						c_p = CO2p.SolidProperties(4,T,P)
						c_v = CO2p.SolidProperties(5,T,P)
						a   = CO2p.SolidProperties(6,T,P)
						dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
					else:
						# Melting line (SLE)
						dfr = pd.DataFrame([i, T, P_d, P_e, "SLE"])
						dfr.style.apply(highlight_max, color='green')
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfr.to_excel(writer, sheet_name="Results", startrow=i+2)
						# Solid properties
						d   = CO2p.SolidProperties(1,T,P)
						s   = CO2p.SolidProperties(2,T,P)
						h   = CO2p.SolidProperties(3,T,P)
						c_p = CO2p.SolidProperties(4,T,P)
						c_v = CO2p.SolidProperties(5,T,P)
						a   = CO2p.SolidProperties(6,T,P)
						dfs = pd.DataFrame([i, T, P_d, d, s, h, c_p, c_v, a])
						dfs.style.apply(highlight_max, color='green')
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfs.to_excel(writer, sheet_name="co2Sproperies", startrow=i+2)
						# Liquid properties
						dl  = density(1,T,P)
						s_l = CO2p.VLProperties(1,T,dl)
						h_l = CO2p.VLProperties(2,T,dl)
						cpl = CO2p.VLProperties(3,T,dl)
						cvl = CO2p.VLProperties(4,T,dl)
						a_l = CO2p.VLProperties(5,T,dl)
						tc  = CO2p.VLProperties(6,T,dl)
						v   = CO2p.VLProperties(7,T,dl)
						st  = CO2p.VLProperties(8,T,dl)
						dfl = pd.DataFrame([i, T, P_d, dl, s_l, h_l, cpl, cvl, a_l, tc, v, st])
						dfl.style.apply(highlight_max, color='green')
						with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
							dfl.to_excel(writer, sheet_name="co2Lproperies", startrow=i+2)                      
				else:
					# Saturation line (VLE)
					dfr = pd.DataFrame([i, T, P_d, P_e, "VLE"])
					dfr.style.apply(highlight_max, color='red')
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfr.to_excel(writer, sheet_name="Results", startrow=i+2)                   
					# Liquid properties
					dl  = density(1,T,P)
					s_l = CO2p.VLProperties(1,T,dl)
					h_l = CO2p.VLProperties(2,T,dl)
					cpl = CO2p.VLProperties(3,T,dl)
					cvl = CO2p.VLProperties(4,T,dl)
					a_l = CO2p.VLProperties(5,T,dl)
					tc  = CO2p.VLProperties(6,T,dl)
					v   = CO2p.VLProperties(7,T,dl)
					st  = CO2p.VLProperties(8,T,dl)
					dfl = pd.DataFrame([i, T, P_d, dl, s_l, h_l, cpl, cvl, a_l, tc, v, st])
					dfl.style.apply(highlight_max, color='red')
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfl.to_excel(writer, sheet_name="co2Lproperies", startrow=i+2)
					# Vapor properties
					dv  = density(2,T,P)
					s_v = CO2p.VLProperties(1,T,dv)
					h_v = CO2p.VLProperties(2,T,dv)
					cpv = CO2p.VLProperties(3,T,dv)
					cvv = CO2p.VLProperties(4,T,dv)
					a_v = CO2p.VLProperties(5,T,dv)
					tc  = CO2p.VLProperties(6,T,dv)
					v   = CO2p.VLProperties(7,T,dv)
					dfv = pd.DataFrame([i, T, P_d, dv, s_v, h_v, cpv, cvv, a_v, tc, v])
					dfv.style.apply(highlight_max, color='red')
					with ExcelWriter(path, mode="a",engine="openpyxl",if_sheet_exists="overlay",) as writer:
						dfv.to_excel(writer, sheet_name="co2Vproperies", startrow=i+2)
	else:
		# Record out of data range
		raise Exception('Data at T =', T,' K and pressure P =',P_d,\
				' MPa out of the admissible range, i.e. 73.15-323.15 K and 0.01-10 MPa') 

# Delete the epty rows within the file
wb = load_workbook('PureCO2.xlsx')
ws1 = wb1['co2Sproperies']
ws2 = wb2['co2Lproperies']
ws3 = wb3['co2Vproperies']
ws4 = wb4['co2Cproperies']
for row in ws1.iter_rows():
	if not any([cell.value for cell in row[1:]]):
		ws1.delete_rows(row[0].row)
for row in ws2.iter_rows():
	if not any([cell.value for cell in row[1:]]):
		ws2.delete_rows(row[0].row)
for row in ws3.iter_rows():
	if not any([cell.value for cell in row[1:]]):
		ws3.delete_rows(row[0].row)
for row in ws4.iter_rows():
	if not any([cell.value for cell in row[1:]]):
		ws4.delete_rows(row[0].row)

# save the excel file
writer.save()

# End of the analysis
print("End of calculation. See the results in PureCO2.xlsx file")
print("If you want to plot them used the CO2plot.py routine")

#--------------------------------------------------------------------------------------------------
# Below the functions used within the algorithm
#--------------------------------------------------------------------------------------------------

def EquilibriumSolver(eps,k,T):
	"""
	#-------------------------------------------
	# Function providing equilibrium pressure)
	# of carbon dioxide at a given temperature
	#-------------------------------------------
	"""
	# Equilibrium phase index: 1 = SVE, 2 = VLE, 3 = SLE
	# Number of iteration
	nmax = 100

	# Set up start value of x in units of Pa
	if k == 1:
		x = SWcE.P_eq(1,T)
	elif k == 2:
		x = SWcE.P_eq(2,T)
	elif k == 3:
		x = SWcE.P_eq(3,T)

	try:
		# Start loop over n
		for n in range(nmax-1):
			# Set up the value of x_old (Value of x at step n-1) and the new value of x
			x_old = x

			#---------------------------------------------
			# Newton-Raphson algorithm used to calculate
			# the zero of a non-analytic function
			#---------------------------------------------
			x = opt.newton(f,x,fprime=df, args=(k,T), tol = eps, maxiter=nmax)

			# Break condition: stop the cycle when |x - x_old|< 10^(-6)
			if mt.abs(x-x_old)<eps: 
				break
	except:
		# Error: if n == nmax the convergence is not achieved
		print('error - convergence not achieved')
	
	# Error: negative pressure
	if x < 0: 
		raise Exception('WARNING: P_eq', x*1.0E-06, 'calculated at temperature', T,'has no physical meaning')

	return x

def f(x,k,T):
	"""
	#-------------------------------------------------------------------
	# This function describes the equation of state used in this model:
	# 1) Span-Wagner and Jäger-Span equation (SLE and SVE)
	#    --> f = RT_0g - RTPhi_r - RTPhi_0 - M(P/(rho))
	# 2) Span-Wagner equation (VLE)
	#    --> f = RT(Phi_r(L)-Phi_r(V) + ln(rho(L)/rho(V)))- BP
	#        with B = M(1/rho(V))-1/rho(L)
	#-------------------------------------------------------------------
	"""

	# Equilibrium phase index: 1 = SVE, 2 = VLE, 3 = SLE

	# JS's molar Gibbs free energy
	g = CO2p.SolidProperties(0,T,x)

	##! What is the purpose of this loop?
	for m in range(1):
		# Set up density
		d[m]   = density(m,T,x)
		# Set up the dimensionless SW's Helmholtz energy
		A_r[m] = CO2p.VLProperties(0,T,d(m))

	#--------------------------------------------------
	# Function f(x)-> with variable x = pressure [Pa]
	#--------------------------------------------------
	if k == 1:
		return (g - R*T*A_r(2) - Mco2*(x/d(2)))
	elif k == 2:
		return (R*T*(A_r(1) - A_r(2)) - Mco2*x*((1/d(2)) - (1/d(1))))
	elif k == 3:
		return g - R*T*A_r(1) - Mco2*(x/d(1))

def df(x,k,T):
	"""
	#----------------------------------------------
	# Function describing the derivative of f(x)
	#----------------------------------------------
	"""

	# Equilibrium phase index: 1 = SVE, 2 = VLE, 3 = SLE

	# Derivative of JS's molar Gibbs free energy
	dg = Mco2/CO2p.SolidProperties(1,T,x)

	#--------------------------------------------------
	# Function f'(x)-> with variable x = pressure [Pa]
	#--------------------------------------------------
	if k == 1:
		return (dg - Mco2*(1/density(2,T,x)))
	elif k == 2:
		return (Mco2*((1/density(1,T,x)) - (1/density(2,T,x))))
	elif k == 3:
		return (dg - Mco2*(1/density(1,T,x)))

def highlight_max(color):
	"""
	#-----------------------------------------------------------
	# Function used to bold and color valuses within excel file
	#-----------------------------------------------------------
	"""

	return ["font-weight: bold", f"color: {color};" ]